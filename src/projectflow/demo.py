from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from projectflow.config import AppConfig, RepertoireChantierConfig, UserConfig
from projectflow.core.fiche_service import FICHE_SUFFIX, FicheService
from projectflow.core.local_repertoire import LocalWorkbookGateway
from projectflow.core.project_service import ProjectService
from projectflow.core.repertoire_service import RepertoireService
from projectflow.outlook.local import create_local_outlook_client
from projectflow.platform.paths import data_dir


@dataclass(slots=True)
class DemoServiceContainer:
    config: AppConfig
    workbook_path: Path
    fiche_service: FicheService | None = None

    def fiche(self) -> FicheService:
        if self.fiche_service is None:
            self.fiche_service = FicheService()
        return self.fiche_service

    def repertoire(self) -> RepertoireService:
        return RepertoireService(LocalWorkbookGateway(self.workbook_path))

    def project(self) -> ProjectService:
        return ProjectService(
            config=self.config,
            fiche_service=self.fiche(),
            repertoire_service=self.repertoire(),
            outlook=create_local_outlook_client(self.config.outlook),
        )


def build_demo_environment(
    *,
    base_dir: Path | None = None,
) -> tuple[AppConfig, DemoServiceContainer]:
    root = base_dir or data_dir() / "demo"
    clients_dir = root / "Clients"
    reference_dir = root / "Modeles" / "10-Racine"
    repertoire_path = root / "Repertoire chantier demo.xlsx"

    clients_dir.mkdir(parents=True, exist_ok=True)
    reference_dir.mkdir(parents=True, exist_ok=True)
    _ensure_reference_fiche(reference_dir)
    _ensure_repertoire(repertoire_path)

    config = AppConfig()
    config.user = UserConfig(
        tenant_id="demo",
        user_id="demo",
        display_name="Mode demo",
        email="demo@projectflow.local",
    )
    config.paths.racine_projets = clients_dir
    config.paths.dossier_reference = reference_dir
    config.paths.repertoire_chantier = RepertoireChantierConfig(
        drive_id="local-demo",
        item_id=str(repertoire_path),
        display_path=str(repertoire_path),
    )
    services = DemoServiceContainer(config=config, workbook_path=repertoire_path)
    return config, services


def _ensure_reference_fiche(reference_dir: Path) -> None:
    fiche_path = reference_dir / f"modele{FICHE_SUFFIX}"
    if fiche_path.exists():
        return
    workbook = Workbook()
    worksheet = cast("Worksheet", workbook.active)
    worksheet.title = "Fiche"
    worksheet["C3"] = ""
    worksheet["D3"] = "Societe : "
    worksheet["D4"] = "Contact : "
    worksheet["D5"] = "Projet : "
    worksheet["D6"] = "Localisation : "
    worksheet["C6"] = ""
    workbook.save(fiche_path)


def _ensure_repertoire(repertoire_path: Path) -> None:
    current_year = date.today().year
    workbook = load_workbook(repertoire_path) if repertoire_path.exists() else Workbook()

    if "Sheet" in workbook.sheetnames and str(current_year) not in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    for year in [current_year, current_year + 1]:
        worksheet_name = str(year)
        if worksheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(worksheet_name)
            worksheet.append([
                "Numero",
                "Date",
                "Societe",
                "Contact",
                "Description",
                "Gere par",
            ])
            for project_id in range(4995, 5005):
                worksheet.append([f"{year}-{project_id}", "", "", "", "", ""])

    workbook.save(repertoire_path)
    workbook.close()
