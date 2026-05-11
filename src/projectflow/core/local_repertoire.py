from __future__ import annotations

from collections.abc import AsyncIterator
from contextlib import AbstractAsyncContextManager, asynccontextmanager
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from projectflow.exceptions import ProjectCreationError


@dataclass(frozen=True, slots=True)
class CellFormat:
    font: Any
    fill: Any
    border: Any
    alignment: Any
    number_format: str
    protection: Any


@dataclass(frozen=True, slots=True)
class RowFormat:
    cell_formats: list[CellFormat]
    height: float | None


class LocalWorkbookGateway:
    def __init__(self, workbook_path: Path) -> None:
        self._workbook_path = workbook_path
        self._workbook: Workbook | None = None

    def session(self) -> AbstractAsyncContextManager[None]:
        return self._session()

    @asynccontextmanager
    async def _session(self) -> AsyncIterator[None]:
        if self._workbook is not None:
            yield
            return
        self._workbook = load_workbook(self._workbook_path)
        try:
            yield
            self._workbook.save(self._workbook_path)
        finally:
            self._workbook.close()
            self._workbook = None

    async def worksheet_exists(self, worksheet_name: str) -> bool:
        return worksheet_name in self._require_workbook().sheetnames

    async def used_range_values(self, worksheet_name: str) -> list[list[Any]]:
        worksheet = self._worksheet(worksheet_name)
        return [
            list(row)
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
                values_only=True,
            )
        ]

    async def update_range_values(
        self,
        worksheet_name: str,
        address: str,
        values: list[list[Any]],
    ) -> None:
        worksheet = self._worksheet(worksheet_name)
        min_col, min_row, _max_col, _max_row = _range_boundaries(address)
        for row_offset, row_values in enumerate(values):
            for column_offset, value in enumerate(row_values):
                cell = worksheet.cell(
                    row=min_row + row_offset,
                    column=min_col + column_offset,
                )
                cell.value = value
                if isinstance(value, date):
                    cell.number_format = "DD.MM.YYYY"

    async def insert_blank_row(
        self,
        worksheet_name: str,
        row_index: int,
        *,
        copy_format_from_row_index: int | None = None,
        format_width: int = 12,
    ) -> None:
        worksheet = self._worksheet(worksheet_name)
        target_row = row_index + 1
        row_format = _capture_row_format(
            worksheet,
            copy_format_from_row_index,
            min_col=1,
            max_col=format_width,
        )
        worksheet.insert_rows(target_row)
        _expand_tables_for_inserted_row(worksheet, target_row)
        if row_format is not None:
            _apply_row_format(worksheet, row_format, target_row=target_row, min_col=1)
        for column_index in range(1, format_width + 1):
            worksheet.cell(row=target_row, column=column_index).value = None

    def _require_workbook(self) -> Workbook:
        if self._workbook is None:
            raise ProjectCreationError("Le repertoire local n'est pas ouvert en session.")
        return self._workbook

    def _worksheet(self, worksheet_name: str) -> Worksheet:
        workbook = self._require_workbook()
        if worksheet_name not in workbook.sheetnames:
            raise ProjectCreationError(f"Onglet introuvable: {worksheet_name}")
        return cast("Worksheet", workbook[worksheet_name])


def _range_boundaries(address: str) -> tuple[int, int, int, int]:
    return cast("tuple[int, int, int, int]", range_boundaries(address))


def _expand_tables_for_inserted_row(worksheet: Worksheet, row_number: int) -> None:
    for table in worksheet.tables.values():
        min_col, min_row, max_col, max_row = _range_boundaries(table.ref)
        if min_row < row_number <= max_row + 1:
            table.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row + 1}"
            )


def _capture_row_format(
    worksheet: Worksheet,
    row_index: int | None,
    *,
    min_col: int,
    max_col: int,
) -> RowFormat | None:
    if row_index is None:
        return None
    source_row = row_index + 1
    if source_row < 1 or source_row > worksheet.max_row:
        return None
    return RowFormat(
        cell_formats=[
            _capture_cell_format(cast("Cell", worksheet.cell(row=source_row, column=column_index)))
            for column_index in range(min_col, max_col + 1)
        ],
        height=worksheet.row_dimensions[source_row].height,
    )


def _apply_row_format(
    worksheet: Worksheet,
    row_format: RowFormat,
    *,
    target_row: int,
    min_col: int,
) -> None:
    for offset, cell_format in enumerate(row_format.cell_formats):
        _apply_cell_format(
            cast("Cell", worksheet.cell(row=target_row, column=min_col + offset)),
            cell_format,
        )
    worksheet.row_dimensions[target_row].height = row_format.height


def _capture_cell_format(cell: Cell) -> CellFormat:
    return CellFormat(
        font=copy(cell.font),
        fill=copy(cell.fill),
        border=copy(cell.border),
        alignment=copy(cell.alignment),
        number_format=cell.number_format,
        protection=copy(cell.protection),
    )


def _apply_cell_format(cell: Cell, cell_format: CellFormat) -> None:
    cell.font = copy(cell_format.font)
    cell.fill = copy(cell_format.fill)
    cell.border = copy(cell_format.border)
    cell.alignment = copy(cell_format.alignment)
    cell.number_format = cell_format.number_format
    cell.protection = copy(cell_format.protection)
