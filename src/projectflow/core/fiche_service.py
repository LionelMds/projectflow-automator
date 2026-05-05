from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Final, cast

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from projectflow.core.models import ProjectInput
from projectflow.core.numero import ProjectNumber
from projectflow.exceptions import ProjectCreationError

FICHE_SUFFIX: Final[str] = " - Fiche dossier clients.xlsx"
PREFIX_RE = re.compile(
    "^\\s*(?P<label>societe|soci\u00e9t\u00e9|contact|projet|localisation)\\s*:\\s*",
    re.IGNORECASE,
)


@dataclass(frozen=True, slots=True)
class FicheData:
    number: str = ""
    societe: str = ""
    contact: str = ""
    designation: str = ""
    localisation: str = ""
    gere_par: str = ""


@dataclass(frozen=True, slots=True)
class FicheCandidate:
    path: Path
    size_bytes: int
    modified_timestamp: float


class FicheService:
    def list_candidates(self, project_dir: Path) -> list[FicheCandidate]:
        candidates: list[FicheCandidate] = []
        for path in project_dir.glob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            stat = path.stat()
            candidates.append(
                FicheCandidate(
                    path=path,
                    size_bytes=stat.st_size,
                    modified_timestamp=stat.st_mtime,
                ),
            )
        return sorted(candidates, key=_candidate_sort_key)

    def locate_fiche(self, project_dir: Path) -> Path:
        candidates = self.list_candidates(project_dir)
        if not candidates:
            raise ProjectCreationError(f"Aucune fiche Excel trouvee dans {project_dir}")
        return candidates[0].path

    def standardize_fiche_name(
        self,
        project_dir: Path,
        number: ProjectNumber,
        *,
        fiche_path: Path | None = None,
    ) -> Path:
        source_path = fiche_path or self.locate_fiche(project_dir)
        standard_path = standard_fiche_path(project_dir, number)
        if source_path == standard_path:
            return source_path
        if standard_path.exists():
            return standard_path
        source_path.rename(standard_path)
        return standard_path

    def fill_fiche(self, project_dir: Path, project: ProjectInput) -> Path:
        fiche_path = self.standardize_fiche_name(project_dir, project.number)

        workbook = load_workbook(fiche_path)
        worksheet = _active_worksheet(workbook)
        worksheet["C3"] = str(project.number)
        _write_prefixed(worksheet, "D3", "Societe", project.societe)
        _write_prefixed(worksheet, "D4", "Contact", project.contact)
        _write_prefixed(worksheet, "D5", "Projet", project.designation)
        _write_prefixed(worksheet, "D6", "Localisation", project.localisation)
        if project.gere_par.strip():
            worksheet["C6"] = project.gere_par.strip()
        workbook.save(fiche_path)
        return fiche_path

    def fill_subproject_fiche(self, project_dir: Path, project: ProjectInput) -> Path:
        if not project.number.is_subproject:
            return self.fill_fiche(project_dir, project)

        target_path = standard_fiche_path(project_dir, project.number)
        if not target_path.exists():
            parent_path = standard_fiche_path(project_dir, project.number.parent)
            source_path = parent_path if parent_path.exists() else self.locate_fiche(project_dir)
            target_path.write_bytes(source_path.read_bytes())

        workbook = load_workbook(target_path)
        worksheet = _active_worksheet(workbook)
        worksheet["C3"] = str(project.number)
        _write_prefixed(worksheet, "D3", "Societe", project.societe)
        _write_prefixed(worksheet, "D4", "Contact", project.contact)
        _write_prefixed(worksheet, "D5", "Projet", project.designation)
        _write_prefixed(worksheet, "D6", "Localisation", project.localisation)
        if project.gere_par.strip():
            worksheet["C6"] = project.gere_par.strip()
        workbook.save(target_path)
        return target_path

    def read_fiche(self, fiche_path: Path) -> FicheData:
        workbook = load_workbook(fiche_path, read_only=True, data_only=True)
        worksheet = _active_worksheet(workbook)
        return FicheData(
            number=_cell_text(worksheet["C3"].value),
            societe=_strip_prefix(_cell_text(worksheet["D3"].value)),
            contact=_strip_prefix(_cell_text(worksheet["D4"].value)),
            designation=_strip_prefix(_cell_text(worksheet["D5"].value)),
            localisation=_strip_prefix(_cell_text(worksheet["D6"].value)),
            gere_par=_cell_text(worksheet["C6"].value),
        )


def standard_fiche_path(project_dir: Path, number: ProjectNumber) -> Path:
    return project_dir / f"{number}{FICHE_SUFFIX}"


def _active_worksheet(workbook: Workbook) -> Worksheet:
    return cast("Worksheet", workbook.active)


def _candidate_sort_key(candidate: FicheCandidate) -> tuple[int, str]:
    contains_fiche = "fiche" in candidate.path.name.lower()
    return (0 if contains_fiche else 1, candidate.path.name.lower())


def _write_prefixed(workbook: Worksheet, cell: str, prefix: str, value: str) -> None:
    if value.strip():
        workbook[cell] = f"{prefix} : {value.strip()}"


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _strip_prefix(value: str) -> str:
    return PREFIX_RE.sub("", value).strip()
