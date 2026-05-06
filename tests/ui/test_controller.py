from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from PySide6.QtWidgets import QMessageBox

from projectflow.config import AppConfig
from projectflow.core.fiche_service import FicheService
from projectflow.core.models import ProjectCreationResult, ProjectInput
from projectflow.core.numero import parse_project_number
from projectflow.core.repertoire_service import NextAvailableProject
from projectflow.ui.controller import ProjectFlowController
from projectflow.ui.main_window import MainWindow


class FakeProjectService:
    def __init__(self) -> None:
        self.created: list[tuple[ProjectInput, bool, bool]] = []
        self.updated: list[ProjectInput] = []
        self.creation_result = ProjectCreationResult(
            project_dir_created=True,
            project_dir="C:/tmp/2026-4995",
            fiche_path="C:/tmp/fiche.xlsx",
        )

    async def create_project(
        self,
        project: ProjectInput,
        *,
        force_overwrite: bool = False,
        update_existing_info: bool = True,
    ) -> ProjectCreationResult:
        self.created.append((project, force_overwrite, update_existing_info))
        return self.creation_result

    async def update_project(self, project: ProjectInput) -> ProjectCreationResult:
        self.updated.append(project)
        return ProjectCreationResult(
            project_dir_created=False,
            project_dir="C:/tmp/2026-4995",
            fiche_path="C:/tmp/fiche.xlsx",
        )


class FakeRepertoireService:
    async def next_available(self, *, year: int) -> NextAvailableProject:
        assert year == 2026
        return NextAvailableProject(number=parse_project_number("2026-4995"), row_index=1)


class FakeServices:
    def __init__(self, fiche_service: FicheService | None = None) -> None:
        self.project_service = FakeProjectService()
        self.repertoire_service = FakeRepertoireService()
        self.fiche_service = fiche_service or FicheService()

    def fiche(self) -> FicheService:
        return self.fiche_service

    def repertoire(self) -> FakeRepertoireService:
        return self.repertoire_service

    def project(self) -> FakeProjectService:
        return self.project_service


def _window(qtbot: Any, tmp_path: Path) -> tuple[MainWindow, AppConfig, FakeServices]:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    window = MainWindow(config)
    qtbot.addWidget(window)
    services = FakeServices()
    ProjectFlowController(window=window, config=config, services=services)  # type: ignore[arg-type]
    window.creation_tab.set_project_identity(year="2026", project_id="4995")
    window.creation_tab.designation_edit.setText("Escalier")
    return window, config, services


@pytest.mark.asyncio
async def test_controller_create_project_reads_form(qtbot: Any, tmp_path: Path) -> None:
    window, _config, services = _window(qtbot, tmp_path)
    controller = ProjectFlowController(
        window=window,
        config=_config,
        services=services,  # type: ignore[arg-type]
    )

    await controller.create_project()

    assert str(services.project_service.created[0][0].number) == "2026-4995"
    assert services.project_service.created[0][0].designation == "Escalier"
    assert services.project_service.created[0][1:] == (False, True)
    assert "Projet cree" in window.creation_tab.logs.toPlainText()


@pytest.mark.asyncio
async def test_controller_logs_outlook_creation_result(qtbot: Any, tmp_path: Path) -> None:
    window, config, services = _window(qtbot, tmp_path)
    config.outlook.enabled = True
    services.project_service.creation_result = ProjectCreationResult(
        project_dir_created=True,
        project_dir="C:/tmp/2026-4995",
        fiche_path="C:/tmp/fiche.xlsx",
        outlook_folder_created=True,
    )
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    await controller.create_project()

    logs = window.creation_tab.logs.toPlainText()
    assert "Dossiers Outlook crees" in logs


@pytest.mark.asyncio
async def test_controller_next_available_prefills_identity(qtbot: Any, tmp_path: Path) -> None:
    window, config, services = _window(qtbot, tmp_path)
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    await controller.next_available()

    assert window.creation_tab.project_id_edit.text() == "4995"
    assert window.creation_tab.subproject_edit.text() == ""


def test_controller_load_project_reads_existing_fiche(qtbot: Any, tmp_path: Path) -> None:
    window, config, services = _window(qtbot, tmp_path)
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["C3"] = "2026-4995"
    worksheet["D3"] = "Societe : Balz"
    worksheet["D5"] = "Projet : Escalier charge"
    workbook.save(project_dir / "2026-4995 - Fiche dossier clients.xlsx")
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    controller.load_project()

    assert window.creation_tab.societe_edit.text() == "Balz"
    assert window.creation_tab.designation_edit.text() == "Escalier charge"


def test_controller_load_project_does_not_rename_fiche(
    qtbot: Any,
    tmp_path: Path,
) -> None:
    window, config, services = _window(qtbot, tmp_path)
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    source_path = project_dir / "ancienne fiche.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["C3"] = "2026-4995"
    worksheet["D5"] = "Projet : Charge sans renommer"
    workbook.save(source_path)
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    controller.load_project()

    assert window.creation_tab.designation_edit.text() == "Charge sans renommer"
    assert source_path.exists()
    assert not (project_dir / "2026-4995 - Fiche dossier clients.xlsx").exists()


@pytest.mark.asyncio
async def test_controller_recreate_existing_project_without_changes_skips_info_update(
    qtbot: Any,
    tmp_path: Path,
) -> None:
    window, config, services = _window(qtbot, tmp_path)
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["C3"] = "2026-4995"
    worksheet["D5"] = "Projet : Escalier"
    workbook.save(project_dir / "2026-4995 - Fiche dossier clients.xlsx")
    services.project_service.creation_result = ProjectCreationResult(
        project_dir_created=False,
        project_dir=str(project_dir),
        fiche_path=str(project_dir / "2026-4995 - Fiche dossier clients.xlsx"),
    )
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    await controller.create_project()

    assert services.project_service.created[0][1:] == (False, False)
    assert "Informations existantes conservees" in window.creation_tab.logs.toPlainText()


@pytest.mark.asyncio
async def test_controller_recreate_existing_project_confirms_changed_information(
    qtbot: Any,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    window, config, services = _window(qtbot, tmp_path)
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["C3"] = "2026-4995"
    worksheet["D5"] = "Projet : Ancien"
    workbook.save(project_dir / "2026-4995 - Fiche dossier clients.xlsx")
    monkeypatch.setattr(
        "PySide6.QtWidgets.QMessageBox.question",
        lambda *_args, **_kwargs: QMessageBox.StandardButton.Yes,
    )
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    await controller.create_project()

    assert services.project_service.created[0][1:] == (True, True)


def test_controller_open_fiche_uses_default_app(
    qtbot: Any,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    window, config, services = _window(qtbot, tmp_path)
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(project_dir / "2026-4995 - Fiche dossier clients.xlsx")
    opened: list[Path] = []
    monkeypatch.setattr(
        "projectflow.ui.controller.open_file_default_app",
        lambda path: opened.append(path) is None or True,
    )
    controller = ProjectFlowController(
        window=window,
        config=config,
        services=services,  # type: ignore[arg-type]
    )

    controller.open_fiche()

    assert opened == [project_dir / "2026-4995 - Fiche dossier clients.xlsx"]
