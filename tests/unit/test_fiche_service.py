from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from projectflow.core.fiche_service import FicheService, standard_fiche_path
from projectflow.core.models import ProjectInput
from projectflow.core.numero import parse_project_number


def test_fill_fiche_prefers_fiche_candidate_and_renames_to_standard(tmp_path: Path) -> None:
    project_dir = tmp_path
    workbook = Workbook()
    path = project_dir / "modele fiche.xlsx"
    workbook.save(path)
    project = ProjectInput(
        number=parse_project_number("2026-4995"),
        designation="Escalier",
        societe="Balz",
        contact="Lionel",
        localisation="Zurich",
        gere_par="LM",
    )

    fiche_path = FicheService(today=lambda: date(2026, 5, 6)).fill_fiche(project_dir, project)

    assert fiche_path == standard_fiche_path(project_dir, project.number)
    loaded = FicheService().read_fiche(fiche_path)
    assert loaded.number == "2026-4995"
    assert loaded.societe == "Balz"
    assert loaded.contact == "Lionel"
    assert loaded.designation == "Escalier"
    assert loaded.localisation == "Zurich"
    assert loaded.gere_par == "LM"
    workbook = load_workbook(fiche_path)
    assert workbook.active["B9"].value.date() == date(2026, 5, 6)
    assert workbook.active["B9"].number_format == "DD.MM.YYYY"


def test_fill_fiche_keeps_existing_creation_date(tmp_path: Path) -> None:
    path = tmp_path / "modele fiche.xlsx"
    workbook = Workbook()
    workbook.active["B9"] = date(2025, 1, 2)
    workbook.save(path)
    project = ProjectInput(number=parse_project_number("2026-4995"))

    fiche_path = FicheService(today=lambda: date(2026, 5, 6)).fill_fiche(tmp_path, project)

    loaded = load_workbook(fiche_path)
    assert loaded.active["B9"].value.date() == date(2025, 1, 2)


def test_read_fiche_strips_prefixes_case_insensitively(tmp_path: Path) -> None:
    path = tmp_path / "fiche.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["C3"] = "2026-4995"
    worksheet["D3"] = "société : Balz"
    worksheet["D4"] = "CONTACT : Lionel"
    worksheet["D5"] = "Projet : Escalier"
    worksheet["D6"] = "Localisation : Zurich"
    worksheet["C6"] = "LM"
    workbook.save(path)

    loaded = FicheService().read_fiche(path)

    assert loaded.societe == "Balz"
    assert loaded.contact == "Lionel"


def test_standardize_fiche_name_renames_selected_file(tmp_path: Path) -> None:
    source = tmp_path / "ancienne fiche.xlsx"
    Workbook().save(source)

    renamed = FicheService().standardize_fiche_name(
        tmp_path,
        parse_project_number("2026-4995"),
        fiche_path=source,
    )

    assert renamed == tmp_path / "2026-4995 - Fiche dossier clients.xlsx"
    assert renamed.exists()
    assert not source.exists()
