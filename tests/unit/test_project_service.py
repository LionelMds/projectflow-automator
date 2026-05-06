from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from projectflow.config import AppConfig, OutlookFolderConfig
from projectflow.core.fiche_service import FicheService
from projectflow.core.models import ProjectInput
from projectflow.core.numero import parse_project_number
from projectflow.core.project_service import (
    ProjectService,
    copy_reference_tree,
    outlook_folder_paths,
    render_outlook_folder_name,
)
from projectflow.exceptions import ConfigError


class FakeRepertoireService:
    def __init__(self) -> None:
        self.calls: list[tuple[ProjectInput, bool]] = []

    async def upsert_project(self, project: ProjectInput, *, force_overwrite: bool = False) -> None:
        self.calls.append((project, force_overwrite))


class FakeOutlook:
    def __init__(self) -> None:
        self.paths: list[list[str]] = []
        self.validated = False

    async def validate_target(self) -> None:
        self.validated = True

    async def ensure_folder_path(self, names: list[str]) -> object:
        self.paths.append(names)
        return object()


class BrokenOutlook:
    async def validate_target(self) -> None:
        raise ConfigError("Compte Outlook introuvable")

    async def ensure_folder_path(self, names: list[str]) -> object:
        del names
        return object()


def test_copy_reference_tree_does_not_overwrite_existing_files(tmp_path: Path) -> None:
    reference = tmp_path / "reference"
    project = tmp_path / "project"
    reference.mkdir()
    project.mkdir()
    (reference / "a.txt").write_text("reference", encoding="utf-8")
    (project / "a.txt").write_text("existing", encoding="utf-8")

    copy_reference_tree(reference, project)

    assert (project / "a.txt").read_text(encoding="utf-8") == "existing"


def test_outlook_folder_templates_render_project_placeholders() -> None:
    project = ProjectInput(number=parse_project_number("2026-4995"), designation="Escalier")
    arborescence = [
        OutlookFolderConfig(
            name="Clients",
            children=[
                OutlookFolderConfig(
                    name="[YYYY]",
                    children=[
                        OutlookFolderConfig(name="[NUMERO] - [XXXX]"),
                    ],
                )
            ],
        ),
    ]

    assert render_outlook_folder_name("[YYYY]-[XXXX]", project) == "2026-4995"
    assert outlook_folder_paths(project, arborescence) == [["Clients", "2026", "2026-4995 - 4995"]]


@pytest.mark.asyncio
async def test_create_project_creates_folder_copies_reference_and_calls_integrations(
    tmp_path: Path,
) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    config.paths.dossier_reference = tmp_path / "reference"
    config.paths.dossier_reference.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(config.paths.dossier_reference / "modele fiche.xlsx")
    config.outlook.enabled = True

    repertoire = FakeRepertoireService()
    outlook = FakeOutlook()
    pinned: list[Path] = []
    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=repertoire,  # type: ignore[arg-type]
        outlook=outlook,
        pin_path=pinned.append,
    )
    project = ProjectInput(number=parse_project_number("2026-4995"), designation="Escalier")

    result = await service.create_project(project)

    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    assert result.project_dir_created is True
    assert Path(result.project_dir) == project_dir
    assert (project_dir / "2026-4995 - Fiche dossier clients.xlsx").exists()
    assert repertoire.calls == [(project, False)]
    assert outlook.validated is True
    assert outlook.paths == [["2026", "2026-4995"]]
    assert pinned == [project_dir]


@pytest.mark.asyncio
async def test_recreate_existing_project_reapplies_integrations_without_updating_info(
    tmp_path: Path,
) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    config.paths.dossier_reference = tmp_path / "reference"
    config.paths.dossier_reference.mkdir(parents=True)
    Workbook().save(config.paths.dossier_reference / "modele fiche.xlsx")
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    fiche_path = project_dir / "2026-4995 - Fiche dossier clients.xlsx"
    Workbook().save(fiche_path)
    config.outlook.enabled = True

    repertoire = FakeRepertoireService()
    outlook = FakeOutlook()
    pinned: list[Path] = []
    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=repertoire,  # type: ignore[arg-type]
        outlook=outlook,
        pin_path=pinned.append,
    )
    project = ProjectInput(number=parse_project_number("2026-4995"), designation="Nouveau texte")

    result = await service.create_project(project, update_existing_info=False)

    assert result.project_dir_created is False
    assert result.fiche_path == str(fiche_path)
    assert not (project_dir / "modele fiche.xlsx").exists()
    assert repertoire.calls == []
    assert outlook.paths == [["2026", "2026-4995"]]
    assert pinned == [project_dir]


@pytest.mark.asyncio
async def test_recreate_existing_project_can_force_information_update(tmp_path: Path) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    config.paths.dossier_reference = tmp_path / "reference"
    config.paths.dossier_reference.mkdir(parents=True)
    Workbook().save(config.paths.dossier_reference / "modele fiche.xlsx")
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    Workbook().save(project_dir / "2026-4995 - Fiche dossier clients.xlsx")
    repertoire = FakeRepertoireService()
    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=repertoire,  # type: ignore[arg-type]
    )
    project = ProjectInput(number=parse_project_number("2026-4995"), designation="Corrige")

    await service.create_project(project, force_overwrite=True, update_existing_info=True)

    assert repertoire.calls == [(project, True)]


@pytest.mark.asyncio
async def test_create_project_skips_outlook_when_disabled(tmp_path: Path) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    config.paths.dossier_reference = tmp_path / "reference"
    config.paths.dossier_reference.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(config.paths.dossier_reference / "modele fiche.xlsx")

    repertoire = FakeRepertoireService()
    outlook = FakeOutlook()
    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=repertoire,  # type: ignore[arg-type]
        outlook=outlook,
    )

    result = await service.create_project(
        ProjectInput(number=parse_project_number("2026-4995"), designation="Escalier"),
    )

    assert result.outlook_folder_created is False
    assert outlook.validated is False
    assert outlook.paths == []


@pytest.mark.asyncio
async def test_create_project_validates_outlook_before_file_operations(tmp_path: Path) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    config.paths.dossier_reference = tmp_path / "reference"
    config.paths.dossier_reference.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(config.paths.dossier_reference / "modele fiche.xlsx")
    config.outlook.enabled = True

    repertoire = FakeRepertoireService()
    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=repertoire,  # type: ignore[arg-type]
        outlook=BrokenOutlook(),
    )

    with pytest.raises(ConfigError, match="Compte Outlook introuvable"):
        await service.create_project(
            ProjectInput(number=parse_project_number("2026-4995"), designation="Escalier"),
        )

    assert not (config.paths.racine_projets / "2026" / "2026-4995").exists()
    assert repertoire.calls == []


@pytest.mark.asyncio
async def test_create_project_requires_local_outlook_connector_when_enabled(tmp_path: Path) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    config.paths.dossier_reference = tmp_path / "reference"
    config.paths.dossier_reference.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(config.paths.dossier_reference / "modele fiche.xlsx")
    config.outlook.enabled = True

    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=FakeRepertoireService(),  # type: ignore[arg-type]
        outlook=None,
    )

    with pytest.raises(ConfigError, match="connecteur Outlook local"):
        await service.create_project(
            ProjectInput(number=parse_project_number("2026-4995"), designation="Escalier"),
        )


@pytest.mark.asyncio
async def test_create_subproject_reuses_parent_folder_without_integrations(tmp_path: Path) -> None:
    config = AppConfig()
    config.paths.racine_projets = tmp_path / "clients"
    project_dir = config.paths.racine_projets / "2026" / "2026-4995"
    project_dir.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(project_dir / "2026-4995 - Fiche dossier clients.xlsx")
    repertoire = FakeRepertoireService()
    outlook = FakeOutlook()
    service = ProjectService(
        config=config,
        fiche_service=FicheService(),
        repertoire_service=repertoire,  # type: ignore[arg-type]
        outlook=outlook,
    )
    project = ProjectInput(number=parse_project_number("2026-4995-2"), designation="Variante")

    result = await service.create_project(project)

    assert Path(result.project_dir) == project_dir
    assert (project_dir / "2026-4995-2 - Fiche dossier clients.xlsx").exists()
    assert repertoire.calls == [(project, False)]
    assert outlook.paths == []
