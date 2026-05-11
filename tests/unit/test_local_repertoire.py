from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table

from projectflow.core.local_repertoire import LocalWorkbookGateway
from projectflow.core.models import ProjectInput
from projectflow.core.numero import parse_project_number
from projectflow.core.repertoire_service import RepertoireService

TODAY = date(2026, 5, 11)


def _create_repertoire(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "2026"
    worksheet.append([
        "Numero",
        "Date",
        "Societe",
        "Contact",
        "Description",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
    ])
    worksheet.append(["2026-4995", "", "", "", "", "", "", "", "", "", "", ""])
    worksheet.append(["2026-5000", "", "", "", "", "", "", "", "", "", "", ""])
    workbook.save(path)
    workbook.close()


def _create_styled_repertoire(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "2026"
    worksheet.append([
        "Numero",
        "Date",
        "Societe",
        "Contact",
        "Description",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
    ])
    worksheet.append([
        "2026-4995",
        TODAY,
        "Balz",
        "Lionel",
        "Escalier",
        "Parent-F",
        "Parent-G",
        "Parent-H",
        "Parent-I",
        "Parent-J",
        "Parent-K",
        "Parent-L",
    ])
    worksheet.append([
        "2026-5000",
        "",
        "",
        "",
        "",
        "Next-F",
        "Next-G",
        "Next-H",
        "Next-I",
        "Next-J",
        "Next-K",
        "Next-L",
    ])
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for column_index in range(1, 13):
        worksheet.cell(row=3, column=column_index).fill = fill
    worksheet.row_dimensions[3].height = 28
    workbook.save(path)
    workbook.close()


@pytest.mark.asyncio
async def test_local_repertoire_updates_main_project(tmp_path: Path) -> None:
    path = tmp_path / "repertoire.xlsx"
    _create_repertoire(path)
    service = RepertoireService(LocalWorkbookGateway(path), today=lambda: TODAY)
    project = ProjectInput(
        number=parse_project_number("2026-4995"),
        designation="Escalier",
        societe="Balz",
        contact="Lionel",
        localisation="Zurich",
        gere_par="LM",
    )

    await service.upsert_project(project)

    workbook = load_workbook(path)
    worksheet = workbook["2026"]
    assert worksheet["A2"].value == "2026-4995"
    assert worksheet["B2"].value.date() == TODAY
    assert worksheet["B2"].number_format == "DD.MM.YYYY"
    assert worksheet["C2"].value == "Balz"
    assert worksheet["D2"].value == "Lionel"
    assert worksheet["E2"].value == "Escalier"
    workbook.close()


@pytest.mark.asyncio
async def test_local_repertoire_inserts_subproject_without_overwriting_next_row(
    tmp_path: Path,
) -> None:
    path = tmp_path / "repertoire.xlsx"
    _create_repertoire(path)
    service = RepertoireService(LocalWorkbookGateway(path), today=lambda: TODAY)
    parent = ProjectInput(number=parse_project_number("2026-4995"), designation="Escalier")
    subproject = ProjectInput(number=parse_project_number("2026-4995-2"), designation="Variante")

    await service.upsert_project(parent)
    await service.upsert_project(subproject)

    workbook = load_workbook(path)
    worksheet = workbook["2026"]
    assert worksheet["A2"].value == "2026-4995"
    assert worksheet["A3"].value == "2026-4995-2"
    assert worksheet["A4"].value == "2026-5000"
    workbook.close()


@pytest.mark.asyncio
async def test_local_repertoire_inserts_subproject_with_blank_values_and_available_row_format(
    tmp_path: Path,
) -> None:
    path = tmp_path / "repertoire.xlsx"
    _create_styled_repertoire(path)
    service = RepertoireService(LocalWorkbookGateway(path), today=lambda: TODAY)
    subproject = ProjectInput(
        number=parse_project_number("2026-4995-2"),
        designation="Variante",
        contact="Contact saisi",
    )

    await service.upsert_project(subproject)

    workbook = load_workbook(path)
    worksheet = workbook["2026"]
    assert worksheet["A3"].value == "2026-4995-2"
    assert worksheet["B3"].value.date() == TODAY
    assert worksheet["B3"].number_format == "DD.MM.YYYY"
    assert worksheet["C3"].value is None
    assert worksheet["D3"].value == "Contact saisi"
    assert worksheet["E3"].value == "Variante"
    assert worksheet["F3"].value is None
    assert worksheet["G3"].value is None
    assert worksheet["L3"].value is None
    assert worksheet["A4"].value == "2026-5000"
    assert worksheet["F2"].value == "Parent-F"
    assert worksheet["L2"].value == "Parent-L"
    assert worksheet["F4"].value == "Next-F"
    assert worksheet["L4"].value == "Next-L"
    assert worksheet["A3"].fill.fgColor.rgb == "00FFF2CC"
    assert worksheet.row_dimensions[3].height == 28
    workbook.close()


@pytest.mark.asyncio
async def test_local_repertoire_inserts_subproject_without_copying_accounting_table_values(
    tmp_path: Path,
) -> None:
    path = tmp_path / "repertoire.xlsx"
    _create_styled_repertoire(path)
    workbook = load_workbook(path)
    worksheet = workbook["2026"]
    worksheet.add_table(Table(displayName="Repertoire", ref="A1:L3"))
    workbook.save(path)
    workbook.close()
    service = RepertoireService(LocalWorkbookGateway(path), today=lambda: TODAY)

    await service.upsert_project(
        ProjectInput(number=parse_project_number("2026-4995-2"), designation="Variante"),
    )

    workbook = load_workbook(path)
    worksheet = workbook["2026"]
    assert worksheet["A3"].value == "2026-4995-2"
    assert worksheet["E3"].value == "Variante"
    assert [worksheet.cell(row=3, column=column).value for column in range(6, 13)] == [
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ]
    assert [worksheet.cell(row=2, column=column).value for column in range(6, 13)] == [
        "Parent-F",
        "Parent-G",
        "Parent-H",
        "Parent-I",
        "Parent-J",
        "Parent-K",
        "Parent-L",
    ]
    assert [worksheet.cell(row=4, column=column).value for column in range(6, 13)] == [
        "Next-F",
        "Next-G",
        "Next-H",
        "Next-I",
        "Next-J",
        "Next-K",
        "Next-L",
    ]
    assert worksheet.tables["Repertoire"].ref == "A1:L4"
    workbook.close()
